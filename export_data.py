import os
import pandas as pd
from openpyxl import load_workbook

def classification_report_export(report_dict, excel_path, sheet_name='Sheet1'):
    # Prepare DataFrame from report_dict
    rows = []
    for key, values in report_dict.items():
        if not isinstance(values, dict):
            continue
        rows.append({
            'Class': key,
            'Precision': values.get('precision', ''),
            'Recall': values.get('recall', ''),
            'F1-Score': values.get('f1-score', ''),
            'Support': values.get('support', '')
        })
    accuracy = report_dict.get('accuracy', None)
    if accuracy is not None:
        rows.append({'Class': 'Accuracy', 'Precision': '', 'Recall': '', 'F1-Score': accuracy, 'Support': ''})
    df = pd.DataFrame(rows)

    if os.path.exists(excel_path):
        # File exists, try appending
        try:
            book = load_workbook(excel_path)
        except Exception as e:
            print(f"Warning: Could not load workbook (corrupted?): {e}")
            print("Creating a new file instead.")
            df.to_excel(excel_path, sheet_name=sheet_name, index=False)
            return

        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            if sheet_name in book.sheetnames:
                startrow = book[sheet_name].max_row + 2  # Add 2 blank rows before appending
            else:
                startrow = 0
            df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=startrow == 0)
    else:
        # File does not exist, create new
        df.to_excel(excel_path, sheet_name=sheet_name, index=False)
